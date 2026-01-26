#!/usr/bin/env python3
"""
TAU2-Bench Evaluation Report Generator
Uses actual TAU2 Pass^k metrics with Excel formulas for all calculations.
"""
import json
import sys
import argparse
from pathlib import Path
from typing import Dict, List, Any
from datetime import datetime
import math
import subprocess
import platform
from json import JSONDecodeError

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing required packages...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "pandas", "openpyxl"])
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.utils import get_column_letter

def _safe_json_loads(s: str) -> dict:
    try:
        return json.loads(s)
    except Exception:
        return {}


def _summarize_request(user_request_raw: str) -> str:
    """
    요청(원문 JSON)에서 핵심만 추출해서 한 셀에 보기 좋게 요약.
    기대 포맷: {reason_for_call, known_info, task_instructions, ...}
    """
    d = _safe_json_loads(user_request_raw or "")
    reason = (d.get("reason_for_call") or "").strip()
    known = (d.get("known_info") or "").strip()
    instr = (d.get("task_instructions") or "").strip()
    # task_instructions는 길어지므로 첫 1~2줄만
    instr_short = ""
    if instr:
        lines = [ln.strip() for ln in instr.splitlines() if ln.strip()]
        instr_short = " / ".join(lines[:2])
        if len(lines) > 2:
            instr_short += " ..."
    parts = []
    if reason:
        parts.append(f"- 문제: {reason}")
    if known:
        parts.append(f"- 정보: {known}")
    if instr_short:
        parts.append(f"- 조건: {instr_short}")
    return "\n".join(parts) if parts else (user_request_raw or "")


def _summarize_gt(gt_raw: str) -> str:
    """
    GT(원문/필수액션)에서 '필수 툴'과 '환경 assertions'를 요약.
    """
    actions = []
    asserts = []
    # 현재 gt_raw는 actions 리스트(JSON 문자열)로 저장됨
    d = _safe_json_loads(gt_raw or "")
    if isinstance(d, list):
        actions = [a.get("name") for a in d if isinstance(a, dict) and a.get("name")]
        for a in d:
            pass
    # env_assertions는 task_meta에 없어서(현재 저장 형태) GT 원문 전체는 숨김 컬럼에서 보게 유도
    if actions:
        return "필수 툴: " + ", ".join(sorted(set(actions)))
    return gt_raw or ""


def _summarize_model(tool_names: str, agent_final: str) -> str:
    """
    모델 결과(툴/최종응답)를 한 셀에 보기 좋게 요약.
    """
    tools = (tool_names or "").strip()
    final = (agent_final or "").strip()
    final_short = final
    if len(final_short) > 240:
        final_short = final_short[:240] + "..."
    parts = []
    parts.append(f"- 툴: {tools if tools else '(없음)'}")
    if final_short:
        parts.append(f"- 최종응답: {final_short}")
    return "\n".join(parts)

def _fmt_kv_call(name: str, args: dict | None) -> str:
    try:
        a = json.dumps(args or {}, ensure_ascii=False)
    except Exception:
        a = str(args)
    return f"{name}({a})"


def _extract_failed_env_assertions(reward_info: dict) -> list[str]:
    out: list[str] = []
    for item in (reward_info or {}).get("env_assertions") or []:
        if not isinstance(item, dict):
            continue
        met = item.get("met")
        env_a = item.get("env_assertion") or {}
        if met is False and isinstance(env_a, dict):
            fn = env_a.get("func_name") or "unknown_assertion"
            args = env_a.get("arguments") or {}
            out.append(_fmt_kv_call(fn, args) + " 미충족")
    return out


def _extract_action_mismatches(reward_info: dict) -> list[str]:
    out: list[str] = []
    for item in (reward_info or {}).get("action_checks") or []:
        if not isinstance(item, dict):
            continue
        if item.get("action_match") is False:
            act = item.get("action") or {}
            if isinstance(act, dict):
                name = act.get("name") or "unknown_action"
                args = act.get("arguments") or {}
                out.append(_fmt_kv_call(name, args) + " 불일치")
    return out


def _normalize_reward_key(k: object) -> str:
    s = str(k)
    # RewardType.ENV_ASSERTION 같은 형태 정리
    if "." in s:
        s = s.split(".")[-1]
    return s.strip()


def _get_rb_value(reward_breakdown: dict, key: str) -> float | None:
    if not isinstance(reward_breakdown, dict):
        return None
    want = key.upper()
    for k, v in reward_breakdown.items():
        kk = _normalize_reward_key(k).upper()
        if kk == want:
            try:
                return float(v)
            except Exception:
                return None
    return None


def _make_fail_reason(
    *,
    pass_flag: int,
    termination: str,
    required_tools: list[str],
    called_tools: list[str],
    missing_tools: list[str],
    failed_env_assertions: list[str],
    action_mismatches: list[str],
    tool_args_err_cnt: int,
    tool_args_err_summary: str,
) -> tuple[str, str, str]:
    """
    반환: (실패분류, 한줄, 상세)
    - PASS면 분류/사유는 간단히
    """
    termination = termination or "n/a"
    req = ", ".join(required_tools) if required_tools else "(없음)"
    called = ", ".join(called_tools) if called_tools else "(없음)"
    miss = ", ".join(missing_tools) if missing_tools else "(없음)"
    env_fail = "; ".join(failed_env_assertions) if failed_env_assertions else "(없음)"
    act_fail = "; ".join(action_mismatches) if action_mismatches else "(없음)"

    if pass_flag == 1:
        return "-", "PASS: reward=1.0", "PASS: 필수 액션/환경 assertion 체크를 모두 통과"

    # 분류
    tag = "Unknown"
    if tool_args_err_cnt > 0:
        tag = "Tool misuse / Schema mismatch"
    elif missing_tools:
        tag = "Tool misuse / Missing required actions"
    elif failed_env_assertions:
        tag = "Reasoning/Planning / Env assertion failed"
    elif "too_many_errors" in termination:
        tag = "Infra/API / Too many errors"
    elif "max_steps" in termination or "max_turns" in termination:
        tag = "Loop/timeout / Max steps"

    # 한 줄
    one = f"FAIL: 종료={termination} / 필수툴={req} / 호출툴={called}"
    extras = []
    if missing_tools:
        extras.append(f"누락툴={miss}")
    if failed_env_assertions:
        extras.append(f"깨진 assertion {len(failed_env_assertions)}개")
    if tool_args_err_cnt > 0:
        extras.append(f"tool args JSON 오류 {tool_args_err_cnt}건")
    if extras:
        one += " / " + " / ".join(extras)

    # 상세(체크리스트)
    detail_lines = [
        f"- 종료사유: {termination}",
        f"- 필수 툴(GT): {req}",
        f"- 호출된 툴(모델): {called}",
        f"- 누락된 툴: {miss}",
        f"- action_checks 불일치: {act_fail}",
        f"- env_assertions 실패: {env_fail}",
    ]
    if tool_args_err_cnt > 0:
        detail_lines.append(f"- tool args JSON 파싱 오류: {tool_args_err_summary}")
    return tag, one, "\n".join(detail_lines)

def setup_styles():
    """가독성 중심(최소 색상, 엑셀 기본 톤) 스타일."""
    grid = "D9D9D9"
    header_fill = "F2F2F2"
    header_fill2 = "E7E6E6"
    pass_fill = "E2F0D9"   # 연한 초록
    fail_fill = "FCE4D6"   # 연한 빨강/주황
    pass_row_fill = "F3FAF1"  # 아주 연한 초록(행 강조)
    fail_row_fill = "FFF4F0"  # 아주 연한 주황(행 강조)
    tool_call_row_fill = "FFF8E1"  # TOOL_CALL(아주 연한 노랑)
    tool_result_row_fill = "F3F3F3"  # TOOL_RESULT(아주 연한 회색)
    top1_fill = "FFF2CC"   # 1위(은은한 골드)
    top2_fill = "DDEBF7"   # 2위(은은한 블루)
    top3_fill = "E7E6E6"   # 3위(은은한 그레이)
    return {
        'title': {
            'font': Font(bold=True, size=14, name="Malgun Gothic"),
            'align': Alignment(horizontal="center", vertical="center")
        },
        'section': {
            'font': Font(bold=True, size=12, name="Malgun Gothic", color="1F4E79"),
            'align': Alignment(horizontal="left", vertical="center")
        },
        'header': {
            'fill': PatternFill(start_color=header_fill, end_color=header_fill, fill_type="solid"),
            'font': Font(bold=True, size=10, name="Malgun Gothic"),
            'align': Alignment(horizontal="center", vertical="center", wrap_text=True),
            'border': Border(
                left=Side(style='thin', color=grid),
                right=Side(style='thin', color=grid),
                top=Side(style='thin', color=grid),
                bottom=Side(style='thin', color=grid)
            ),
        },
        'header2': {
            'fill': PatternFill(start_color=header_fill2, end_color=header_fill2, fill_type="solid"),
            'font': Font(bold=True, size=10, name="Malgun Gothic"),
            'align': Alignment(horizontal="center", vertical="center", wrap_text=True),
            'border': Border(
                left=Side(style='thin', color=grid),
                right=Side(style='thin', color=grid),
                top=Side(style='thin', color=grid),
                bottom=Side(style='thin', color=grid)
            ),
        },
        'data': {
            'font': Font(size=9, name="Malgun Gothic"),
            'align': Alignment(horizontal="left", vertical="top", wrap_text=True),
            'border': Border(
                left=Side(style='thin', color=grid),
                right=Side(style='thin', color=grid),
                top=Side(style='thin', color=grid),
                bottom=Side(style='thin', color=grid)
            )
        },
        'data_center': {
            'font': Font(size=9, name="Malgun Gothic"),
            'align': Alignment(horizontal="center", vertical="center"),
            'border': Border(
                left=Side(style='thin', color=grid),
                right=Side(style='thin', color=grid),
                top=Side(style='thin', color=grid),
                bottom=Side(style='thin', color=grid)
            )
        },
        'pass': {
            'fill': PatternFill(start_color=pass_fill, end_color=pass_fill, fill_type="solid"),
            'font': Font(size=9, name="Malgun Gothic", color="006100")
        },
        'fail': {
            'fill': PatternFill(start_color=fail_fill, end_color=fail_fill, fill_type="solid"),
            'font': Font(size=9, name="Malgun Gothic", color="9C0006")
        },
        'pass_row': {
            'fill': PatternFill(start_color=pass_row_fill, end_color=pass_row_fill, fill_type="solid"),
        },
        'fail_row': {
            'fill': PatternFill(start_color=fail_row_fill, end_color=fail_row_fill, fill_type="solid"),
        },
        'tool_call_row': {
            'fill': PatternFill(start_color=tool_call_row_fill, end_color=tool_call_row_fill, fill_type="solid"),
        },
        'tool_result_row': {
            'fill': PatternFill(start_color=tool_result_row_fill, end_color=tool_result_row_fill, fill_type="solid"),
        },
        'fail_strong_font': {
            'font': Font(bold=True, size=9, name="Malgun Gothic", color="9C0006"),
        },
        'top1': {
            'fill': PatternFill(start_color=top1_fill, end_color=top1_fill, fill_type="solid"),
            'font': Font(bold=True, size=10, name="Malgun Gothic"),
        },
        'top2': {
            'fill': PatternFill(start_color=top2_fill, end_color=top2_fill, fill_type="solid"),
            'font': Font(bold=True, size=9, name="Malgun Gothic"),
        },
        'top3': {
            'fill': PatternFill(start_color=top3_fill, end_color=top3_fill, fill_type="solid"),
            'font': Font(bold=True, size=9, name="Malgun Gothic"),
        },
    }

def _extract_tool_args_json_errors(messages: list[dict]) -> tuple[int, str, str]:
    """
    messages[*].raw_data에 포함된 원본 tool_call.function.arguments 문자열을 기준으로
    JSON 파싱 실패(빈 문자열/깨진 JSON)를 감지한다.

    NOTE: 런타임에서는 llm_utils.py에서 크래시를 막기 위해 arguments를 빈 dict로 대체할 수 있으므로,
    '파싱 실패 여부'는 여기에서 raw_data를 기준으로 판정해야 한다.
    """
    errs: list[dict] = []

    for m in messages or []:
        if (m.get("role") or "") != "assistant":
            continue
        raw = m.get("raw_data") or {}
        # LiteLLM choice.to_dict() 형태: {"message": {...}, ...}
        raw_msg = raw.get("message") if isinstance(raw, dict) else None
        if not isinstance(raw_msg, dict):
            continue
        tool_calls = raw_msg.get("tool_calls") or []
        if not isinstance(tool_calls, list):
            continue
        for tc in tool_calls:
            if not isinstance(tc, dict):
                continue
            func = tc.get("function") or {}
            if not isinstance(func, dict):
                continue
            name = func.get("name") or ""
            args = func.get("arguments")
            if args is None:
                errs.append({"tool": name, "error": "arguments is None", "raw": None})
                continue
            if isinstance(args, dict):
                # 이미 dict면 OK
                continue
            s = str(args).strip()
            if s == "" or s.lower() in {"null", "none"}:
                errs.append({"tool": name, "error": "arguments empty", "raw": s})
                continue
            try:
                json.loads(s)
            except JSONDecodeError as e:
                errs.append(
                    {
                        "tool": name,
                        "error": f"invalid JSON ({e.msg})",
                        "raw": s[:500],
                    }
                )
            except Exception as e:
                errs.append(
                    {
                        "tool": name,
                        "error": f"invalid JSON ({type(e).__name__})",
                        "raw": s[:500],
                    }
                )

    if not errs:
        return 0, "", ""

    # 사람용 요약
    summary_parts = []
    for e in errs[:6]:
        t = e.get("tool") or "unknown_tool"
        msg = e.get("error") or "error"
        summary_parts.append(f"{t}: {msg}")
    summary = "; ".join(summary_parts)
    if len(errs) > 6:
        summary += f" (+{len(errs)-6} more)"

    return len(errs), summary, json.dumps(errs, ensure_ascii=False)


def create_runs_raw_sheet(wb, runs, styles):
    """런 단위 원본 데이터(요청/GT/모델응답 포함). 모든 집계는 이 시트를 기반으로 파생."""
    ws = wb.create_sheet("런_원본", 0)
    
    # Headers
    headers = [
        "RunID",
        "모델",
        "LLM(Agent)",
        "도메인",
        "도메인 설명",
        "TaskID",
        "Trial",
        "결과",
        "성공(0/1)",
        "Reward",
        "RewardBreakdown(JSON)",
        "종료사유",
        "툴호출수",
        "툴목록(요약)",
        "툴호출(JSON 원본)",
        "툴응답(원본)",
        "요청(원본)",
        "GT(원본/필수액션)",
        "모델 최종응답(원본)",
    ]
    ws.append(headers)
    
    for col_idx, cell in enumerate(ws[1], 1):
        cell.font = styles['header']['font']
        cell.fill = styles['header']['fill']
        cell.alignment = styles['header']['align']
        cell.border = styles['header']['border']
    
    # Add all run data
    for run in runs:
        row_idx = ws.max_row + 1
        ws.append(
            [
                run["RunID"],
                run["ModelLabel"],
                run["AgentLLM"],
                run["Domain"],
                None,  # 도메인 설명(수식)
                run["TaskID"],
                run["Trial"],
                None,  # 결과(수식)
                run["Pass"],
                run["Reward"],
                run.get("RewardBreakdownJSON", ""),
                run.get("Termination", "N/A"),
                run.get("ToolCallCount", 0),
                run.get("ToolNames", ""),
                run.get("ToolCallsRaw", ""),
                run.get("ToolResultsRaw", ""),
                run.get("UserRequestRaw", ""),
                run.get("GTRaw", ""),
                run.get("AgentFinalRaw", ""),
            ]
        )
        # 도메인 설명 수식
        ws.cell(row=row_idx, column=5).value = f'=IFERROR(VLOOKUP(D{row_idx},도메인_설명!$A:$C,3,FALSE),"")'
        # 결과(PASS/FAIL)는 엑셀 수식으로
        ws.cell(row=row_idx, column=8).value = f'=IF(I{row_idx}=1,"PASS","FAIL")'
    
    # Style data rows
    for row_idx in range(2, ws.max_row + 1):
        # PASS/FAIL만 아주 옅게 강조 (결과 컬럼)
        result_cell = ws.cell(row=row_idx, column=8)
        pass_flag = ws.cell(row=row_idx, column=9).value
        if pass_flag == 1:
            result_cell.fill = styles["pass"]["fill"]
            result_cell.font = styles["pass"]["font"]
        else:
            result_cell.fill = styles["fail"]["fill"]
            result_cell.font = styles["fail"]["font"]

        for col_idx in range(1, 20):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = styles['data']['border']
            if col_idx in [6, 7, 9, 10, 13]:
                cell.alignment = styles['data_center']['align']
            elif col_idx == 8:
                cell.alignment = styles['data_center']['align']
            else:
                cell.alignment = styles['data']['align']
            
            if col_idx == 10:  # Reward column
                cell.number_format = '0.0000'
    
    # Freeze header and add filters
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    
    # Column widths
    ws.column_dimensions['A'].width = 34  # RunID
    ws.column_dimensions['B'].width = 28  # 모델
    ws.column_dimensions['C'].width = 28  # LLM
    ws.column_dimensions['D'].width = 10  # 도메인
    ws.column_dimensions['E'].width = 24  # 도메인 설명
    ws.column_dimensions['F'].width = 8   # TaskID
    ws.column_dimensions['G'].width = 6   # Trial
    ws.column_dimensions['H'].width = 8   # 결과
    ws.column_dimensions['I'].width = 9   # 성공
    ws.column_dimensions['J'].width = 9   # Reward
    ws.column_dimensions['K'].width = 28  # reward breakdown
    ws.column_dimensions['L'].width = 14  # 종료
    ws.column_dimensions['M'].width = 8   # tool count
    ws.column_dimensions['N'].width = 22  # tool names
    ws.column_dimensions['O'].width = 44  # tool calls raw
    ws.column_dimensions['P'].width = 44  # tool results raw
    ws.column_dimensions['Q'].width = 48  # request raw
    ws.column_dimensions['R'].width = 48  # GT raw
    ws.column_dimensions['S'].width = 48  # agent final

    # 보기 좋은 행높이
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 60

def create_task_summary_sheet(wb, all_logs, models_mapping, domains, styles):
    """Create task-level summary with Pass^k calculations using Excel formulas."""
    ws = wb.create_sheet("Task별_집계", 1)
    
    # Title
    ws.append(["Task별 성공률 집계 (Pass^k 계산용)"])
    ws.merge_cells('A1:H1')
    ws['A1'].font = styles['title']['font']
    ws['A1'].alignment = styles['title']['align']
    ws.row_dimensions[1].height = 25
    
    ws.append([""])
    ws.append(["이 시트는 각 Task의 시행별 성공 횟수를 집계하여 Pass^k 메트릭을 계산합니다."])
    ws.merge_cells('A3:H3')
    ws.append(["Pass^k = COMBIN(성공횟수, k) / COMBIN(총시행횟수, k)"])
    ws.merge_cells('A4:H4')
    ws.append([""])
    
    # Headers (총시행/성공횟수 모두 런_원본 기반 수식)
    headers = ["모델", "도메인", "도메인 설명", "TaskID", "총시행", "성공횟수", "Pass@1", "Pass@2", "Pass@4"]
    ws.append(headers)
    header_row = ws.max_row
    
    for col_idx, cell in enumerate(ws[header_row], 1):
        cell.font = styles['header']['font']
        cell.fill = styles['header']['fill']
        cell.alignment = styles['header']['align']
        cell.border = styles['header']['border']
    
    # Group data by Model, Domain, TaskID (행 생성용 키만 파이썬으로 추출)
    task_groups = {}
    for log in all_logs:
        key = (log['Model'], log['Domain'], str(log['TaskID']))
        if key not in task_groups:
            task_groups[key] = []
        task_groups[key].append(log)
    
    # Add data rows
    for (model, domain, task_id), logs in sorted(task_groups.items()):
        row_num = ws.max_row + 1
        ws.append([model, domain, None, task_id, None, None, None, None, None])

        # 도메인 설명(도메인_설명 시트 참조)
        ws.cell(row=row_num, column=3).value = f'=IFERROR(VLOOKUP(B{row_num},도메인_설명!$A:$C,3,FALSE),"")'

        # 총시행/성공횟수: 런 시트 기반 (외부/연결 오탐 방지)
        # 런: B 모델, C 도메인, D TaskID, I PASS?
        ws.cell(row=row_num, column=5).value = f'=COUNTIFS(런!$B:$B,$A{row_num},런!$C:$C,$B{row_num},런!$D:$D,$D{row_num})'
        ws.cell(row=row_num, column=6).value = f'=COUNTIFS(런!$B:$B,$A{row_num},런!$C:$C,$B{row_num},런!$D:$D,$D{row_num},런!$I:$I,\"PASS\")'
        
        # Pass@1: n>=1일 때만 의미. n=0이면 빈칸(집계에서 제외)
        ws.cell(row=row_num, column=7).value = f"=IF(E{row_num}<1,\"\",IFERROR(F{row_num}/E{row_num},\"\"))"
        ws.cell(row=row_num, column=7).number_format = '0.0%'
        
        # Pass@2: n<2면 '표본부족'으로 계산 불가 → 빈칸(0으로 오해 방지)
        ws.cell(row=row_num, column=8).value = f"=IF(E{row_num}<2,\"\",IFERROR(COMBIN(F{row_num},2)/COMBIN(E{row_num},2),\"\"))"
        ws.cell(row=row_num, column=8).number_format = '0.0%'
        
        # Pass@4: n<4면 '표본부족'으로 계산 불가 → 빈칸(0으로 오해 방지)
        ws.cell(row=row_num, column=9).value = f"=IF(E{row_num}<4,\"\",IFERROR(COMBIN(F{row_num},4)/COMBIN(E{row_num},4),\"\"))"
        ws.cell(row=row_num, column=9).number_format = '0.0%'
        
        # Apply styles
        for col_idx in range(1, 10):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.border = styles['data']['border']
            if col_idx in [4, 5, 6]:
                cell.alignment = styles['data_center']['align']
            elif col_idx >= 7:
                cell.alignment = styles['data_center']['align']
            else:
                cell.alignment = styles['data']['align']
    
    # Freeze header and add filters
    ws.freeze_panes = f"A{header_row+1}"
    ws.auto_filter.ref = f"A{header_row}:I{ws.max_row}"
    
    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    for col in ['G', 'H', 'I']:
        ws.column_dimensions[col].width = 12

def create_summary_sheet(wb, models_mapping, domains, styles):
    """요약 시트: Overall 랭킹 + 도메인별 Pass^k 매트릭스를 한 시트에 섹션으로 구성."""
    ws = wb.create_sheet("요약", 0)
    
    # Title
    ws.append(["TAU2-Bench 평가 요약"])
    ws.merge_cells('A1:O1')
    ws['A1'].font = styles['title']['font']
    ws['A1'].alignment = styles['title']['align']
    ws.row_dimensions[1].height = 25
    
    ws.append([""])
    
    # Description
    desc = "Pass^k = COMBIN(성공횟수, k) / COMBIN(총시행횟수, k) 의 Task 평균. 성공 기준: Reward가 1.0(±1e-6)에 해당."
    ws.append([desc])
    ws.merge_cells('A3:O3')
    ws['A3'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 30

    # ===== 실패사유/체크축 Glossary =====
    ws.append([""])
    ws.append(["실패사유(종료사유) / 평가체크(Reward 축) 빠른 해석 가이드"])
    gloss_title_row = ws.max_row
    ws.merge_cells(f"A{gloss_title_row}:O{gloss_title_row}")
    ws.cell(row=gloss_title_row, column=1).font = styles["section"]["font"]
    ws.row_dimensions[gloss_title_row].height = 20

    ws.append(
        [
            "읽는 순서: (1) termination_reason(왜 멈췄나) → (2) reward_info(어떤 체크를 못 맞췄나). "
            "AGENT_STOP/USER_STOP만 '정상 평가'로 들어가며, 그 외(MAX_STEPS/TOO_MANY_ERRORS/AGENT_ERROR/USER_ERROR)는 조기종료로 reward=0 처리됩니다."
        ]
    )
    gloss_desc_row = ws.max_row
    ws.merge_cells(f"A{gloss_desc_row}:O{gloss_desc_row}")
    ws.cell(row=gloss_desc_row, column=1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[gloss_desc_row].height = 42

    # termination_reason table
    ws.append(["구분", "값(termination_reason)", "의미(직관)", "자주 보이는 실패 상황/해석"])
    tr_h = ws.max_row
    for c in range(1, 5):
        cell = ws.cell(tr_h, c)
        cell.font = styles["header"]["font"]
        cell.fill = styles["header"]["fill"]
        cell.alignment = styles["header"]["align"]
        cell.border = styles["header"]["border"]

    tr_rows = [
        ["정상종료(가능)", "agent_stop", "에이전트가 종료 토큰으로 종료", "종료는 했어도 ACTION/ENV/DB 등을 못 맞추면 FAIL"],
        ["정상종료(가능)", "user_stop", "유저가 STOP/TRANSFER/OUT-OF-SCOPE 등으로 종료", "유저가 포기/전환해도 user_stop로 종료 가능 → 결과는 reward로 판단"],
        ["강제종료", "max_steps", "최대 스텝 초과", "루프/진전 없음 → 조기종료로 reward=0"],
        ["강제종료", "too_many_errors", "오류 누적(툴/프로토콜/예외)", "툴 호출 실패/예외 누적 → 조기종료로 reward=0"],
        ["오류종료", "agent_error", "에이전트 프로토콜 위반/예외", "빈 메시지, 텍스트+툴콜 혼합 등 → 조기종료로 reward=0"],
        ["오류종료", "user_error", "유저 프로토콜 위반/예외", "유저 시뮬레이터 규칙 위반 → 조기종료로 reward=0"],
    ]
    for rr in tr_rows:
        ws.append(rr)
        r = ws.max_row
        for c in range(1, 5):
            cell = ws.cell(r, c)
            cell.border = styles["data"]["border"]
            cell.alignment = styles["data"]["align"] if c != 2 else styles["data_center"]["align"]

    ws.append([""])

    # reward axis table
    ws.append(["축", "데이터(리포트에서 보는 곳)", "의미(직관)", "FAIL이면 바로 보는 포인트"])
    ax_h = ws.max_row
    for c in range(1, 5):
        cell = ws.cell(ax_h, c)
        cell.font = styles["header"]["font"]
        cell.fill = styles["header"]["fill"]
        cell.alignment = styles["header"]["align"]
        cell.border = styles["header"]["border"]

    ax_rows = [
        ["ACTION", "런: 필수툴/호출툴/누락툴 + action_checks 불일치", "필수 툴/행동(절차) 수행 여부", "누락툴이 있는지, action_match=false 항목"],
        ["ENV_ASSERTION", "런: 깨진 env_assertions", "최종 환경 조건(assertion) 만족 여부", "met=false인 assertion 목록(예: speed=excellent)"],
        ["DB", "RewardBreakdown/db_check(숨김 컬럼)", "골드 vs 예측 DB 상태 일치 여부", "db_match=false (도메인에 따라)"],
        ["COMMUNICATE", "RewardBreakdown/communicate_checks(도메인에 따라)", "사용자 커뮤니케이션 요구사항 충족", "communicate가 0 또는 note 확인"],
        ["NL_ASSERTION", "RewardBreakdown/nl_assertions(있는 경우)", "자연어 assertion 충족(WIP 포함)", "실패한 assertion 항목"],
    ]
    for rr in ax_rows:
        ws.append(rr)
        r = ws.max_row
        for c in range(1, 5):
            cell = ws.cell(r, c)
            cell.border = styles["data"]["border"]
            cell.alignment = styles["data"]["align"]

    # 폭 조정(Glossary 영역)
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 8, 12)
    ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width or 40, 26)
    ws.column_dimensions["C"].width = max(ws.column_dimensions["C"].width or 14, 26)
    ws.column_dimensions["D"].width = max(ws.column_dimensions["D"].width or 14, 54)

    # ===== 집계 흐름(스코어 계산) 다이어그램 =====
    ws.append([""])
    ws.append(["스코어 집계 흐름(Trial → Task → Domain → Overall)"])
    flow_title_row = ws.max_row
    ws.merge_cells(f"A{flow_title_row}:O{flow_title_row}")
    ws.cell(row=flow_title_row, column=1).font = styles["section"]["font"]

    flow_lines = [
        "1) Trial(대화 1회) → PASS(0/1): reward=1.0이면 PASS, 아니면 FAIL",
        "2) Task(같은 Task를 n번 반복) → Pass^k: COMBIN(성공횟수 c, k) / COMBIN(총시행 n, k)",
        "3) Domain → 해당 도메인의 Task들의 Pass^k 평균(매크로 평균)",
        "4) Overall → 도메인(Retail/Airline/Telecom) 점수의 평균(동일 가중치)",
        "엑셀 기준: 런(원천) → Task별_집계(숨김, Pass^k 계산) → 요약(도메인/Overall 평균)",
    ]
    ws.append(["\n".join(flow_lines)])
    flow_desc_row = ws.max_row
    ws.merge_cells(f"A{flow_desc_row}:O{flow_desc_row}")
    ws.cell(row=flow_desc_row, column=1).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[flow_desc_row].height = 84

    # ===== 비용 절감 관점: P@1(=Pass^1)만으로 quick 스크리닝하기 =====
    ws.append([""])
    ws.append(["비용 절감 Quick 기준(권장): P@1만으로 1차 스크리닝 → 상위 모델만 n≥2/4로 재검증"])
    qs_row = ws.max_row
    ws.merge_cells(f"A{qs_row}:O{qs_row}")
    ws.cell(row=qs_row, column=1).font = styles["section"]["font"]

    txt = (
        "현재 리포트처럼 num_trials=1이면 Task당 n=1이라 P@2/P@4는 정의상 계산 불가(n<k)입니다.\n"
        "- 비용 절감 목적의 1차 비교는 P@1(=성공률)에 집중해도 됩니다(큰 격차/회귀 탐지에 유용).\n"
        "- 단, 툴콜/멀티턴 태스크는 분산이 크므로 P@1만으로 '안정성' 결론을 내리면 위험합니다.\n"
        "- 실무 추천: 상위 1~2개 모델만 num_trials=2 또는 4로 추가 실행해 P@2/P@4로 재현성(안정성)을 확인하세요."
    )
    ws.append([txt])
    r = ws.max_row
    ws.merge_cells(f"A{r}:O{r}")
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[r].height = 90

    # ===== 케이스 1개로 보는 스코어가 찍히는 방식(직관) =====
    ws.append([""])
    ws.append(["케이스 예시(런 시트 첫 행 기준) — 멀티턴 왕복 → 체크 → Reward/PASS"])
    ex_title_row = ws.max_row
    ws.merge_cells(f"A{ex_title_row}:O{ex_title_row}")
    ws.cell(row=ex_title_row, column=1).font = styles["section"]["font"]

    ws.append(["항목", "값(런! 첫 케이스에서 자동 참조)", "해석"])
    ex_h = ws.max_row
    for c in range(1, 4):
        cell = ws.cell(ex_h, c)
        cell.font = styles["header"]["font"]
        cell.fill = styles["header"]["fill"]
        cell.alignment = styles["header"]["align"]
        cell.border = styles["header"]["border"]

    # 런 시트: 헤더가 3행, 첫 데이터는 4행으로 가정
    r = ws.max_row
    rows = [
        ("종료사유", "=런!F4", "agent_stop/user_stop이면 정상 평가, 그 외는 조기종료로 reward=0"),
        ("조기종료?", "=런!G4", "Y면 조기종료(대개 max_steps/too_many_errors/agent_error/user_error)"),
        ("RewardBasis", "=런!J4", "어떤 축이 점수에 포함됐는지(AND처럼 곱해짐)"),
        ("RB_ENV_ASSERTION", "=런!K4", "환경 assertion(예: speed=excellent)이 0이면 실패"),
        ("RB_ACTION", "=런!L4", "필수툴/행동이 0이면 실패"),
        ("누락툴", "=런!R4", "필수툴(GT) 중 실제 호출툴에 없는 것"),
        ("깨진 env_assertions", "=런!S4", "met=false인 assertion 목록(왜 안 됐는지)"),
        ("PASS?", "=런!I4", "reward==1.0이고 조기종료가 아니면 PASS"),
    ]
    for label, formula, explain in rows:
        ws.append([label, formula, explain])
        rr = ws.max_row
        ws.cell(rr, 2).value = formula
        for c in range(1, 4):
            cell = ws.cell(rr, c)
            cell.border = styles["data"]["border"]
            cell.alignment = styles["data"]["align"] if c != 2 else styles["data"]["align"]

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 54
    ws.column_dimensions["C"].width = 54
    
    ws.append([""])
    
    # Section title (glossary가 추가되었으므로 위치가 유동적)
    ws.append(["Overall Pass^k 랭킹 (전 도메인 평균)"])
    rank_title_row = ws.max_row
    ws.merge_cells(f"A{rank_title_row}:F{rank_title_row}")
    ws.cell(row=rank_title_row, column=1).font = styles['section']['font']
    ws.row_dimensions[rank_title_row].height = 20
    
    # Headers
    headers = ["순위", "모델", "Pass@1", "Pass@2", "Pass@4", "RankKey(hidden)"]
    ws.append(headers)
    header_row = ws.max_row
    
    for col_idx, cell in enumerate(ws[header_row], 1):
        cell.font = styles['header']['font']
        cell.fill = styles['header']['fill']
        cell.alignment = styles['header']['align']
        cell.border = styles['header']['border']
    
    # Data rows with formulas (정렬은 엑셀에서 수행. 순위는 RANK 계열 대신 COUNTIF로 호환성 확보)
    first_data_row = ws.max_row + 1
    for _, (model_key, model_name) in enumerate(models_mapping.items(), 1):
        row_num = ws.max_row + 1
        ws.append([None, model_name, None, None, None, None])
        
        # Pass@1: Average of Pass@1 for this model from Task별_집계
        ws.cell(row=row_num, column=3).value = f'=IFERROR(AVERAGEIF(Task별_집계!A:A, B{row_num}, Task별_집계!G:G),"")'
        ws.cell(row=row_num, column=3).number_format = '0.00%'
        
        # Pass@2: Average of Pass@2 for this model
        ws.cell(row=row_num, column=4).value = f'=IFERROR(AVERAGEIF(Task별_집계!A:A, B{row_num}, Task별_집계!H:H),"")'
        ws.cell(row=row_num, column=4).number_format = '0.00%'
        
        # Pass@4: Average of Pass@4 for this model
        ws.cell(row=row_num, column=5).value = f'=IFERROR(AVERAGEIF(Task별_집계!A:A, B{row_num}, Task별_집계!I:I),"")'
        ws.cell(row=row_num, column=5).number_format = '0.00%'
        # RankKey: Pass@1 > Pass@2 > Pass@4 우선, 동점은 행번호로 안정화
        # NOTE: 빈칸/텍스트가 섞여도 랭킹이 깨지지 않도록 N()로 숫자 강제
        ws.cell(row=row_num, column=6).value = (
            f"=N(C{row_num})*1000000 + N(D{row_num})*1000 + N(E{row_num}) + ROW()/1000000000"
        )

        # Apply styles (이 행 전체)
        for col_idx in range(1, 6 + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.border = styles['data']['border']
            if col_idx == 1 or col_idx >= 3:
                cell.alignment = styles['data_center']['align']
            else:
                cell.alignment = styles['data']['align']

    last_data_row = ws.max_row
    # 순위 수식 입력
    # NOTE: COUNTIF("> "&F{r})는 F가 텍스트/빈값 취급되는 케이스에서 모두 같은 순위로 깨질 수 있어
    #       SUMPRODUCT 비교식으로 안전하게 랭킹을 계산한다.
    for r in range(first_data_row, last_data_row + 1):
        ws.cell(row=r, column=1).value = (
            f"=IF(F{r}=\"\",\"\",1+SUMPRODUCT(--($F${first_data_row}:$F${last_data_row}>F{r})))"
        )
        ws.cell(row=r, column=1).alignment = styles["data_center"]["align"]

    # ===== 랭킹 강조(과하지 않게 1/2/3위만) =====
    rank_range = f"A{first_data_row}:E{last_data_row}"
    ws.conditional_formatting.add(
        rank_range,
        FormulaRule(formula=[f"$A{first_data_row}=1"], fill=styles["top1"]["fill"], font=styles["top1"]["font"], stopIfTrue=True),
    )
    ws.conditional_formatting.add(
        rank_range,
        FormulaRule(formula=[f"$A{first_data_row}=2"], fill=styles["top2"]["fill"], font=styles["top2"]["font"], stopIfTrue=True),
    )
    ws.conditional_formatting.add(
        rank_range,
        FormulaRule(formula=[f"$A{first_data_row}=3"], fill=styles["top3"]["fill"], font=styles["top3"]["font"], stopIfTrue=True),
    )

    # RankKey 컬럼 숨김
    ws.column_dimensions["F"].hidden = True
    
    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14

    # ===== Tool-call JSON args health section =====
    ws.append([""])
    ws.append(["툴콜 arguments(JSON) 안정성(=Schema mismatch 후보)"])
    health_title_row = ws.max_row
    ws.merge_cells(f"A{health_title_row}:F{health_title_row}")
    ws.cell(row=health_title_row, column=1).font = styles["section"]["font"]

    ws.append(
        [
            "정의: assistant tool_calls의 원본 raw_data에서 function.arguments가 빈 문자열/깨진 JSON이면 오류로 집계(실패를 완화하지 않고, 실패 원인을 투명하게 비교하기 위한 지표)."
        ]
    )
    health_desc_row = ws.max_row
    ws.merge_cells(f"A{health_desc_row}:O{health_desc_row}")
    ws.cell(row=health_desc_row, column=1).alignment = Alignment(
        horizontal="left", vertical="center", wrap_text=True
    )
    ws.row_dimensions[health_desc_row].height = 36

    ws.append(["모델", "런 수", "오류 런 수", "오류율", "런당 평균 오류수", "비고"])
    health_header_row = ws.max_row
    for col_idx, cell in enumerate(ws[health_header_row], 1):
        cell.font = styles["header"]["font"]
        cell.fill = styles["header"]["fill"]
        cell.alignment = styles["header"]["align"]
        cell.border = styles["header"]["border"]

    for model_name in models_mapping.values():
        r = ws.max_row + 1
        ws.append([model_name, None, None, None, None, ""])
        # 런 시트 기준: B=모델, T=ToolArgsJSONErrorCount(hidden)
        ws.cell(r, 2).value = f'=COUNTIF(런!$B:$B, $A{r})'
        ws.cell(r, 3).value = f'=COUNTIFS(런!$B:$B, $A{r}, 런!$T:$T, ">0")'
        ws.cell(r, 4).value = f"=IFERROR(C{r}/B{r},0)"
        ws.cell(r, 4).number_format = "0.00%"
        ws.cell(r, 5).value = f'=IFERROR(AVERAGEIFS(런!$T:$T, 런!$B:$B, $A{r}),0)'
        ws.cell(r, 5).number_format = "0.00"
        for cc in range(1, 6 + 1):
            cell = ws.cell(r, cc)
            cell.border = styles["data"]["border"]
            cell.alignment = (
                styles["data_center"]["align"]
                if cc in [2, 3, 4, 5]
                else styles["data"]["align"]
            )

    # ===== Domain matrix section =====
    ws.append([""])
    ws.append(["모델 × 도메인 Pass^k 매트릭스"])
    matrix_title_row = ws.max_row
    ws.merge_cells(f"A{matrix_title_row}:O{matrix_title_row}")
    ws.cell(row=matrix_title_row, column=1).font = styles["section"]["font"]

    ws.append([f"평가 일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ts_row = ws.max_row
    ws.merge_cells(f"A{ts_row}:O{ts_row}")

    ws.append(["각 도메인별 Pass^k 값(= 각 Task의 Pass^k 평균)."])
    desc_row = ws.max_row
    ws.merge_cells(f"A{desc_row}:O{desc_row}")

    ws.append([""])
    header_row_1 = ["도메인"]
    for model_name in models_mapping.values():
        header_row_1.extend([model_name, "", ""])
    ws.append(header_row_1)
    h1 = ws.max_row
    header_row_2 = [""]
    for _ in models_mapping.values():
        header_row_2.extend(["P@1", "P@2", "P@4"])
    ws.append(header_row_2)
    h2 = ws.max_row

    # merge + style headers
    ws.merge_cells(start_row=h1, start_column=1, end_row=h2, end_column=1)
    for rr in [h1, h2]:
        for cc in range(1, 1 + len(header_row_2)):
            cell = ws.cell(rr, cc)
            cell.border = styles["header"]["border"]
            cell.alignment = styles["header"]["align"]
            if rr == h1:
                cell.fill = styles["header"]["fill"]
                cell.font = styles["header"]["font"]
            else:
                cell.fill = styles["header2"]["fill"]
                cell.font = styles["header2"]["font"]

    col = 2
    for _mn in models_mapping.values():
        ws.merge_cells(start_row=h1, start_column=col, end_row=h1, end_column=col + 2)
        col += 3

    domain_names = {"retail": "Retail", "airline": "Airline", "telecom": "Telecom"}
    data_start = ws.max_row + 1
    for d in domains:
        r = ws.max_row + 1
        ws.cell(r, 1).value = domain_names.get(d, d)
        ws.cell(r, 1).border = styles["data"]["border"]
        ws.cell(r, 1).alignment = styles["data"]["align"]
        col = 2
        for _k, model_name in models_mapping.items():
            ws.cell(r, col).value = f'=IFERROR(AVERAGEIFS(Task별_집계!G:G, Task별_집계!A:A, \"{models_mapping[_k]}\", Task별_집계!B:B, \"{d}\"),"")'
            ws.cell(r, col).number_format = "0.00%"
            ws.cell(r, col+1).value = f'=IFERROR(AVERAGEIFS(Task별_집계!H:H, Task별_집계!A:A, \"{models_mapping[_k]}\", Task별_집계!B:B, \"{d}\"),"")'
            ws.cell(r, col+1).number_format = "0.00%"
            ws.cell(r, col+2).value = f'=IFERROR(AVERAGEIFS(Task별_집계!I:I, Task별_집계!A:A, \"{models_mapping[_k]}\", Task별_집계!B:B, \"{d}\"),"")'
            ws.cell(r, col+2).number_format = "0.00%"
            for cc in [col, col+1, col+2]:
                c = ws.cell(r, cc)
                c.border = styles["data"]["border"]
                c.alignment = styles["data_center"]["align"]
            col += 3

    overall_r = ws.max_row + 1
    ws.cell(overall_r, 1).value = "Overall"
    ws.cell(overall_r, 1).font = Font(bold=True, size=10, name="Malgun Gothic")
    ws.cell(overall_r, 1).fill = styles["header2"]["fill"]
    ws.cell(overall_r, 1).alignment = styles["data_center"]["align"]
    ws.cell(overall_r, 1).border = styles["data"]["border"]
    data_end = overall_r - 1
    col = 2
    for _ in models_mapping.values():
        ws.cell(overall_r, col).value = f"=IFERROR(AVERAGE({get_column_letter(col)}{data_start}:{get_column_letter(col)}{data_end}),\"\")"
        ws.cell(overall_r, col).number_format = "0.00%"
        ws.cell(overall_r, col+1).value = f"=IFERROR(AVERAGE({get_column_letter(col+1)}{data_start}:{get_column_letter(col+1)}{data_end}),\"\")"
        ws.cell(overall_r, col+1).number_format = "0.00%"
        ws.cell(overall_r, col+2).value = f"=IFERROR(AVERAGE({get_column_letter(col+2)}{data_start}:{get_column_letter(col+2)}{data_end}),\"\")"
        ws.cell(overall_r, col+2).number_format = "0.00%"
        for cc in [col, col+1, col+2]:
            c = ws.cell(overall_r, cc)
            c.font = Font(bold=True, size=10, name="Malgun Gothic")
            c.fill = styles["header2"]["fill"]
            c.border = styles["data"]["border"]
            c.alignment = styles["data_center"]["align"]
        col += 3

    # ===== 매트릭스 강조(과하지 않게): 각 도메인 행에서 P@1 최고값만 은은하게 표시 =====
    # P@1 컬럼들: 2,5,8,... (모델당 3칸 중 첫번째)
    p1_cols = [2 + 3 * i for i in range(len(models_mapping))]
    for rr in range(data_start, overall_r + 1):
        # 데이터가 없으면(빈칸만) 하이라이트하지 않음
        p1_cells = ",".join([f"{get_column_letter(c)}{rr}" for c in p1_cols])
        for c in p1_cols:
            cell_addr = f"{get_column_letter(c)}{rr}"
            ws.conditional_formatting.add(
                cell_addr,
                FormulaRule(
                    formula=[f"=AND({cell_addr}<>\"\",{cell_addr}=MAX({p1_cells}))"],
                    fill=styles["top1"]["fill"],
                    stopIfTrue=False,
                ),
            )

    # Freeze header for the sheet top (ranking)
    ws.freeze_panes = f"A{header_row+1}"


def create_runs_sheet(wb, runs, styles):
    """
    런 시트(간결): 케이스_요약 + 런_원본을 합친 형태.
    - 기본은 간결한 컬럼만 노출
    - 원본/JSON/툴응답은 숨김 컬럼으로 유지(사용자가 필요시 펼치기)
    """
    ws = wb.create_sheet("런", 1)
    ws.append(["Run 단위 케이스 (스코어/실패원인 먼저 → 필요 시 원문 펼치기)"])
    ws.merge_cells("A1:U1")
    ws["A1"].font = styles["title"]["font"]
    ws["A1"].alignment = styles["title"]["align"]
    ws.row_dimensions[1].height = 22

    ws.append(["읽는 법: (1) 종료사유/조기종료? (2) RewardBasis/RB_* 중 0인 축 (3) 누락툴/깨진 assertion 확인. 원문(JSON/툴로그)은 숨김 컬럼을 펼치면 됩니다."])
    ws.merge_cells("A2:U2")
    ws["A2"].alignment = styles["data"]["align"]
    ws.row_dimensions[2].height = 32

    # Pass@k 관련 주의(표본 부족이면 0으로 보일 수 있음)
    ws.append(["주의: Pass@2/Pass@4는 해당 Task의 총시행(n)이 각각 2/4 미만이면 '계산 불가'라서 0으로 표시될 수 있습니다(FAIL 반영이 안 된 게 아니라 표본 부족)."])
    ws.merge_cells("A3:U3")
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 28

    # 정보 과다 방지: '스코어/실패원인'을 앞에 배치하고, 원문은 숨김 컬럼으로 이동
    headers = [
        "RunID", "모델", "도메인", "TaskID", "Trial",
        "종료사유", "조기종료?", "Reward", "PASS?",
        "RewardBasis",
        "RB_ENV_ASSERTION", "RB_ACTION", "RB_DB", "RB_COMMUNICATE", "RB_NL_ASSERTION",
        "필수툴(GT)", "호출툴(모델)", "누락툴",
        "깨진 env_assertions", "action_checks 불일치",
        "실패분류(L1/L2)",
    ]
    hidden_headers = [
        "결과(PASS/FAIL)",
        "사용자 첫 발화(원문)",
        "요청(원문: 시나리오 JSON)",
        "GT(원문 JSON)",
        "모델 tool_calls(원문)",
        "툴결과(원문)",
        "모델 최종응답(원문)",
        "왜 맞/틀(한줄)",
        "왜 맞/틀(상세)",
        "RewardBreakdown(JSON)",
        "ToolArgsJSONErrorCount",
        "ToolArgsJSONErrorSummary",
        "ToolArgsJSONErrors(JSON)",
    ]
    ws.append(headers + hidden_headers)
    hrow = ws.max_row
    for c in ws[hrow]:
        c.font = styles["header"]["font"]
        c.fill = styles["header"]["fill"]
        c.alignment = styles["header"]["align"]
        c.border = styles["header"]["border"]

    for run in runs:
        rb = run.get("RewardBreakdown") or {}
        tool_args_err_cnt = int(run.get("ToolArgsJSONErrorCount") or 0)
        tool_args_err_summary = run.get("ToolArgsJSONErrorSummary") or ""
        term = str(run.get("Termination") or "")
        req_raw = run.get("UserScenarioRaw","") or run.get("UserRequestRaw","")
        first_user_raw = run.get("UserFirstUtterance","")
        gt_raw = run.get("GTRaw","")
        agent_final_raw = run.get("AgentFinalRaw","")
        tool_calls_raw = run.get("ToolCallsRaw","")
        tool_results_raw = run.get("ToolResultsRaw","")
        required_tools = run.get("RequiredTools") or []
        called_tools = run.get("CalledTools") or []
        missing_tools = run.get("MissingTools") or []
        failed_env_assertions = run.get("FailedEnvAssertions") or []
        action_mismatches = run.get("ActionMismatches") or []

        fail_tag, why_one, why_detail = _make_fail_reason(
            pass_flag=int(run.get("Pass") or 0),
            termination=term,
            required_tools=required_tools,
            called_tools=called_tools,
            missing_tools=missing_tools,
            failed_env_assertions=failed_env_assertions,
            action_mismatches=action_mismatches,
            tool_args_err_cnt=tool_args_err_cnt,
            tool_args_err_summary=tool_args_err_summary,
        )

        row = [
            run.get("RunID",""),
            run.get("ModelLabel",""),
            run.get("Domain",""),
            run.get("TaskID",""),
            run.get("Trial",0),
            term,
            "Y" if term not in {"agent_stop", "user_stop"} else "N",
            run.get("Reward",0.0),
            "PASS" if run.get("Pass")==1 else "FAIL",
            run.get("RewardBasisRaw",""),
            run.get("RB_ENV_ASSERTION"),
            run.get("RB_ACTION"),
            run.get("RB_DB"),
            run.get("RB_COMMUNICATE"),
            run.get("RB_NL_ASSERTION"),
            ", ".join(required_tools) if required_tools else "",
            ", ".join(called_tools) if called_tools else "",
            ", ".join(missing_tools) if missing_tools else "",
            "\n".join(failed_env_assertions) if failed_env_assertions else "",
            "\n".join(action_mismatches) if action_mismatches else "",
            fail_tag,
            # hidden
            "PASS" if run.get("Pass")==1 else "FAIL",
            first_user_raw,
            req_raw,
            gt_raw,
            tool_calls_raw,
            tool_results_raw,
            agent_final_raw,
            why_one,
            why_detail,
            run.get("RewardBreakdownJSON",""),
            run.get("ToolArgsJSONErrorCount", 0),
            tool_args_err_summary,
            run.get("ToolArgsJSONErrorsRaw", ""),
        ]
        ws.append(row)
        r = ws.max_row
        # 스타일
        for col_idx in range(1, len(headers) + len(hidden_headers) + 1):
            cell = ws.cell(r, col_idx)
            cell.border = styles["data"]["border"]
            if col_idx in [5,6,7,8,9]:
                cell.alignment = styles["data_center"]["align"]
            elif col_idx in [11,12,13,14,15]:
                cell.alignment = styles["data_center"]["align"]
            else:
                cell.alignment = styles["data"]["align"]
        # 결과 색
        rc = ws.cell(r, 9)  # PASS?
        if run.get("Pass")==1:
            rc.fill = styles["pass"]["fill"]; rc.font = styles["pass"]["font"]
        else:
            rc.fill = styles["fail"]["fill"]; rc.font = styles["fail"]["font"]
        ws.row_dimensions[r].height = 84

    # ===== PASS/FAIL 행 강조(과하지 않게, 연한 배경) =====
    first_data_row = hrow + 1
    last_data_row = ws.max_row
    # 보이는 핵심 영역(A~U)만 연하게 칠함(숨김 컬럼까지 칠하면 지저분해 보일 수 있음)
    vis_range = f"A{first_data_row}:U{last_data_row}"
    # PASS 행
    ws.conditional_formatting.add(
        vis_range,
        FormulaRule(
            formula=[f'$I{first_data_row}="PASS"'],
            fill=styles["pass_row"]["fill"],
            stopIfTrue=False,
        ),
    )
    # FAIL 행
    ws.conditional_formatting.add(
        vis_range,
        FormulaRule(
            formula=[f'$I{first_data_row}="FAIL"'],
            fill=styles["fail_row"]["fill"],
            stopIfTrue=False,
        ),
    )
    # FAIL이면 "종료사유(F)" + "실패분류(U)"를 조금 더 눈에 띄게
    fail_focus_range = f"F{first_data_row}:F{last_data_row}"
    ws.conditional_formatting.add(
        fail_focus_range,
        FormulaRule(
            formula=[f'$I{first_data_row}="FAIL"'],
            font=styles["fail_strong_font"]["font"],
            stopIfTrue=False,
        ),
    )
    fail_focus_range2 = f"U{first_data_row}:U{last_data_row}"
    ws.conditional_formatting.add(
        fail_focus_range2,
        FormulaRule(
            formula=[f'$I{first_data_row}="FAIL"'],
            font=styles["fail_strong_font"]["font"],
            stopIfTrue=False,
        ),
    )

    ws.freeze_panes = f"A{hrow+1}"
    ws.auto_filter.ref = f"A{hrow}:{get_column_letter(len(headers)+len(hidden_headers))}{ws.max_row}"

    # Column widths (핵심만 보이게)
    widths = {
        "A":34, "B":22, "C":10, "D":10, "E":6,
        "F":12, "G":10, "H":8, "I":8,
        "J":26,
        "K":14, "L":10, "M":10, "N":14, "O":14,
        "P":22, "Q":22, "R":22,
        "S":36, "T":36,
        "U":26,
    }
    for k,v in widths.items():
        ws.column_dimensions[k].width = v
    # 숨김 컬럼(원문/디버깅): headers 다음부터 전부 숨김
    start_hidden = len(headers) + 1
    end_hidden = len(headers) + len(hidden_headers)
    for idx in range(start_hidden, end_hidden + 1):
        ws.column_dimensions[get_column_letter(idx)].hidden = True
    return ws


def create_turns_sheet(wb, turns_rows, styles):
    """턴 단위(대화 흐름을 직관적으로 보는 시트)."""
    ws = wb.create_sheet("대화", 2)
    ws.append(["대화 흐름(원문) + 툴콜/툴결과를 한눈에"])
    ws.merge_cells("A1:N1")
    ws["A1"].font = styles["title"]["font"]
    ws["A1"].alignment = styles["title"]["align"]
    ws.row_dimensions[1].height = 22

    ws.append(["이 시트 의미: 한 줄 = 대화 이벤트 1개(사용자/모델 발화 또는 TOOL_CALL/TOOL_RESULT). RunID/TurnIdx로 정렬하면 '한 케이스의 흐름'이 그대로 보입니다."])
    ws.merge_cells("A2:N2")
    ws["A2"].alignment = styles["data"]["align"]
    ws.row_dimensions[2].height = 28

    ws.append(["색상: TOOL_CALL=연노랑 / TOOL_RESULT=연회색. 사용법: (1) RunID 필터 → (2) TurnIdx 오름차순 → (3) Kind=TOOL_* 행에서 ToolName/Args/Result 확인"])
    ws.merge_cells("A3:N3")
    ws["A3"].alignment = styles["data"]["align"]
    ws.row_dimensions[3].height = 28

    headers = [
        "RunID",
        "모델",
        "도메인",
        "TaskID",
        "Trial",
        "PASS?",
        "TurnIdx",
        "Role",
        "Kind",
        "ToolName(요약)",
        "ToolArgs(원문)",
        "Text(원문)",
        "ToolResult(원문)",
        "ToolCalls(JSON 원문)",
    ]
    ws.append(headers)
    header_row = ws.max_row
    for c in ws[header_row]:
        c.font = styles["header"]["font"]
        c.fill = styles["header"]["fill"]
        c.alignment = styles["header"]["align"]
        c.border = styles["header"]["border"]
    for row in turns_rows:
        ws.append(row)
    for r in range(header_row + 1, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(r, c)
            cell.border = styles["data"]["border"]
            # Trial/PASS?/TurnIdx/Role/Kind는 가운데 정렬
            if c in [5, 6, 7, 8, 9]:
                cell.alignment = styles["data_center"]["align"]
            else:
                cell.alignment = styles["data"]["align"]
        ws.row_dimensions[r].height = 54

    # ===== TOOL_CALL / TOOL_RESULT 행 강조(조건부 서식) =====
    first_data_row = header_row + 1
    last_data_row = ws.max_row
    # ToolCalls(JSON 원문) 컬럼은 숨김이므로, 보이는 범위까지만
    vis_range = f"A{first_data_row}:M{last_data_row}"
    ws.conditional_formatting.add(
        vis_range,
        FormulaRule(
            formula=[f'$I{first_data_row}="TOOL_CALL"'],
            fill=styles["tool_call_row"]["fill"],
            stopIfTrue=False,
        ),
    )
    ws.conditional_formatting.add(
        vis_range,
        FormulaRule(
            formula=[f'$I{first_data_row}="TOOL_RESULT"'],
            fill=styles["tool_result_row"]["fill"],
            stopIfTrue=False,
        ),
    )

    # ===== PASS/FAIL 표시(전용 컬럼만 은은하게) =====
    pass_col_range = f"F{first_data_row}:F{last_data_row}"
    ws.conditional_formatting.add(
        pass_col_range,
        FormulaRule(
            formula=[f'$F{first_data_row}="PASS"'],
            fill=styles["pass"]["fill"],
            font=styles["pass"]["font"],
            stopIfTrue=False,
        ),
    )
    ws.conditional_formatting.add(
        pass_col_range,
        FormulaRule(
            formula=[f'$F{first_data_row}="FAIL"'],
            fill=styles["fail"]["fill"],
            font=styles["fail"]["font"],
            stopIfTrue=False,
        ),
    )

    ws.freeze_panes = f"A{header_row+1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{ws.max_row}"

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 6
    ws.column_dimensions["F"].width = 7   # PASS?
    ws.column_dimensions["G"].width = 7   # TurnIdx
    ws.column_dimensions["H"].width = 10  # Role
    ws.column_dimensions["I"].width = 12  # Kind
    ws.column_dimensions["J"].width = 24  # ToolName
    ws.column_dimensions["K"].width = 44  # ToolArgs
    ws.column_dimensions["L"].width = 54  # Text
    ws.column_dimensions["M"].width = 54  # ToolResult
    ws.column_dimensions["N"].width = 50  # ToolCalls(JSON)
    # ToolCalls(JSON)은 필요할 때만 펼치기
    ws.column_dimensions["N"].hidden = True
    return ws
    
    # Title
    ws.append(["모델 × 도메인 Pass^k 매트릭스"])
    ws.merge_cells(f'A1:{get_column_letter(len(models_mapping) * 3 + 1)}1')
    ws['A1'].font = styles['title']['font']
    ws['A1'].alignment = styles['title']['align']
    ws.row_dimensions[1].height = 25
    
    ws.append([""])
    
    # Timestamp
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([f"평가 일시: {timestamp}"])
    ws.merge_cells(f'A3:{get_column_letter(len(models_mapping) * 3 + 1)}3')
    
    # Description
    ws.append(["각 도메인별 Pass^k 값. Pass^k = 각 Task의 Pass^k를 평균한 값. (Task별_집계 시트 참조)"])
    ws.merge_cells(f'A4:{get_column_letter(len(models_mapping) * 3 + 1)}4')
    
    ws.append([""])
    
    # 2단 헤더: (모델명 merged) + (P@1/P@2/P@4)
    header_row_1 = ["도메인"]
    for model_name in models_mapping.values():
        header_row_1.extend([model_name, "", ""])
    ws.append(header_row_1)
    header_row_idx_1 = ws.max_row

    header_row_2 = [""]
    for _ in models_mapping.values():
        header_row_2.extend(["P@1", "P@2", "P@4"])
    ws.append(header_row_2)
    header_row_idx_2 = ws.max_row

    # 스타일 + merge
    ws.cell(row=header_row_idx_1, column=1).font = styles["header"]["font"]
    ws.cell(row=header_row_idx_1, column=1).fill = styles["header"]["fill"]
    ws.cell(row=header_row_idx_1, column=1).alignment = styles["header"]["align"]
    ws.cell(row=header_row_idx_1, column=1).border = styles["header"]["border"]
    ws.merge_cells(start_row=header_row_idx_1, start_column=1, end_row=header_row_idx_2, end_column=1)

    col = 2
    for _model_name in models_mapping.values():
        ws.merge_cells(start_row=header_row_idx_1, start_column=col, end_row=header_row_idx_1, end_column=col + 2)
        for c in range(col, col + 3):
            cell_top = ws.cell(row=header_row_idx_1, column=c)
            cell_top.font = styles["header"]["font"]
            cell_top.fill = styles["header"]["fill"]
            cell_top.alignment = styles["header"]["align"]
            cell_top.border = styles["header"]["border"]
        for c in range(col, col + 3):
            cell_sub = ws.cell(row=header_row_idx_2, column=c)
            cell_sub.value = header_row_2[c - 1]
            cell_sub.font = styles["header2"]["font"]
            cell_sub.fill = styles["header2"]["fill"]
            cell_sub.alignment = styles["header2"]["align"]
            cell_sub.border = styles["header2"]["border"]
        col += 3
    
    # Domain descriptions
    domain_names = {
        'retail': 'Retail',
        'airline': 'Airline',
        'telecom': 'Telecom'
    }
    
    # Data rows for each domain
    # data rows start
    data_start_row = ws.max_row + 1
    for domain_key in domains:
        row_num = ws.max_row + 1
        ws.cell(row=row_num, column=1).value = domain_names[domain_key]
        ws.cell(row=row_num, column=1).alignment = styles["data"]["align"]
        ws.cell(row=row_num, column=1).border = styles["data"]["border"]

        col = 2
        for _model_key, model_name in models_mapping.items():
            ws.cell(row=row_num, column=col).value = f'=IFERROR(AVERAGEIFS(Task별_집계!F:F, Task별_집계!A:A, "{model_name}", Task별_집계!B:B, "{domain_key}"),0)'
            ws.cell(row=row_num, column=col).number_format = '0.00%'
            ws.cell(row=row_num, column=col + 1).value = f'=IFERROR(AVERAGEIFS(Task별_집계!G:G, Task별_집계!A:A, "{model_name}", Task별_집계!B:B, "{domain_key}"),0)'
            ws.cell(row=row_num, column=col + 1).number_format = '0.00%'
            ws.cell(row=row_num, column=col + 2).value = f'=IFERROR(AVERAGEIFS(Task별_집계!H:H, Task별_집계!A:A, "{model_name}", Task별_집계!B:B, "{domain_key}"),0)'
            ws.cell(row=row_num, column=col + 2).number_format = '0.00%'

            for c in range(col, col + 3):
                cell = ws.cell(row=row_num, column=c)
                cell.alignment = styles["data_center"]["align"]
                cell.border = styles["data"]["border"]
            col += 3
    
    # Overall row
    overall_row = ws.max_row + 1
    ws.cell(row=overall_row, column=1).value = "Overall"
    ws.cell(row=overall_row, column=1).font = Font(bold=True, size=10, name="Malgun Gothic")
    ws.cell(row=overall_row, column=1).fill = styles["header2"]["fill"]
    ws.cell(row=overall_row, column=1).alignment = styles["data_center"]["align"]
    ws.cell(row=overall_row, column=1).border = styles["data"]["border"]

    # data rows range
    data_end_row = overall_row - 1
    col = 2
    for _model_key, _model_name in models_mapping.items():
        ws.cell(row=overall_row, column=col).value = f"=AVERAGE({get_column_letter(col)}{data_start_row}:{get_column_letter(col)}{data_end_row})"
        ws.cell(row=overall_row, column=col).number_format = '0.00%'
        ws.cell(row=overall_row, column=col + 1).value = f"=AVERAGE({get_column_letter(col+1)}{data_start_row}:{get_column_letter(col+1)}{data_end_row})"
        ws.cell(row=overall_row, column=col + 1).number_format = '0.00%'
        ws.cell(row=overall_row, column=col + 2).value = f"=AVERAGE({get_column_letter(col+2)}{data_start_row}:{get_column_letter(col+2)}{data_end_row})"
        ws.cell(row=overall_row, column=col + 2).number_format = '0.00%'
        for c in range(col, col + 3):
            cell = ws.cell(row=overall_row, column=c)
            cell.font = Font(bold=True, size=10, name="Malgun Gothic")
            cell.fill = styles["header2"]["fill"]
            cell.alignment = styles["data_center"]["align"]
            cell.border = styles["data"]["border"]
        col += 3

    # Freeze + filter
    ws.freeze_panes = "B" + str(data_start_row)
    last_col = 1 + (len(models_mapping) * 3)
    ws.auto_filter.ref = f"A{header_row_idx_2}:{get_column_letter(last_col)}{data_end_row}"

    # Column widths
    ws.column_dimensions["A"].width = 12
    for c in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 9

def create_detailed_run_log(wb, all_logs, all_turns, gt_map, styles):
    """(deprecated) kept for backward compatibility; no longer used when using 3-sheet layout."""
    ws = wb.create_sheet("상세_런_로그", 99)
    
    # Headers
    headers = [
        "결과", "모델", "도메인", "도메인 설명", "Query ID", "질문",
        "GT(원본/필수액션)", "툴호출수", "툴목록(원본)",
        "모델 최종응답(원본)", "실패분류(L1/L2)", "근거(원본)", "Reward"
    ]
    
    ws.append(headers)
    for col_idx, cell in enumerate(ws[1], 1):
        cell.font = styles['header']['font']
        cell.fill = styles['header']['fill']
        cell.alignment = styles['header']['align']
        cell.border = styles['header']['border']
    
    # Group turns by run_id
    turn_by_run = {}
    for turn in all_turns:
        run_id = turn['RunID']
        if run_id not in turn_by_run:
            turn_by_run[run_id] = []
        turn_by_run[run_id].append(turn)
    
    # 모든 run 포함 (FAIL 먼저 보고 싶으면 엑셀 필터/정렬)
    for log in all_logs:
        run_id = f"{log['Model']}_{log['Domain']}_T{log['TaskID']}_trial{log['Trial']}"
        turns = turn_by_run.get(run_id, [])
        
        user_question = ""
        for turn in turns:
            if turn['Speaker'] == 'user':
                user_question = turn['Message'][:100] + "..." if len(turn['Message']) > 100 else turn['Message']
                break
        
        # run 전체를 훑어서 tool_calls를 누적
        tool_names: list[str] = []
        tool_count = 0
        for turn in turns:
            tc = turn.get("Tool_Called") or ""
            if tc:
                chunks = [x.strip() for x in tc.split(";") if x.strip()]
                tool_count += len(chunks)
                for ch in chunks:
                    name = ch.split("(")[0].strip()
                    if name:
                        tool_names.append(name)

        # 모델 최종 응답(assistant 마지막 메시지 원본)
        agent_final = ""
        for turn in reversed(turns):
            if turn.get("Speaker") == "assistant":
                agent_final = turn.get("Message") or ""
                break
        
        term_reason = log['Termination']
        # 실패분류 개선(termination만이 아니라 action_checks/env/db/communicate 기반)
        fail_l1l2 = ""
        evidence = ""
        if log.get("Pass") == 1:
            result = "PASS"
            fail_l1l2 = "-"
            evidence = "reward=1.0"
        else:
            result = "FAIL"
            if 'too_many_errors' in (term_reason or ""):
                fail_l1l2 = "Infra/API / Too many errors"
                evidence = term_reason
            elif 'max_turns' in (term_reason or ""):
                fail_l1l2 = "Loop/timeout / Max turns"
                evidence = term_reason
            else:
                # reward breakdown 기반(없으면 termination)
                rb = log.get("RewardBreakdown") or {}
                if rb.get("ACTION") == 0.0:
                    fail_l1l2 = "Tool misuse / Missing required actions"
                    evidence = "ACTION=0"
                elif rb.get("DB") == 0.0:
                    fail_l1l2 = "Reasoning/Planning / DB mismatch"
                    evidence = "DB=0"
                elif rb.get("COMMUNICATE") == 0.0:
                    fail_l1l2 = "Missing info / Communication"
                    evidence = "COMMUNICATE=0"
                else:
                    fail_l1l2 = "Unknown"
                    evidence = term_reason or "n/a"

        gt = gt_map.get((log["Domain"], str(log["TaskID"])), "N/A")
        row_data = [
            result,
            log['Model'],
            log['Domain'],
            None,  # 도메인 설명(수식)
            f"{log['Domain']}_{log['TaskID']}",
            user_question,
            gt,
            tool_count,
            ", ".join(tool_names) if tool_names else "",
            agent_final,
            fail_l1l2,
            evidence,
            log.get("Reward", 0.0),
        ]

        ws.append(row_data)
        row_idx = ws.max_row

        # 도메인 설명 수식
        ws.cell(row=row_idx, column=4).value = f'=IFERROR(VLOOKUP(C{row_idx},도메인_설명!$A:$C,3,FALSE),"")'
        # Reward 표시 포맷
        ws.cell(row=row_idx, column=13).number_format = '0.0000'

        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = styles['data']['border']
            if col_idx == 1:
                if result == "PASS":
                    cell.font = styles['pass']['font']
                    cell.fill = styles['pass']['fill']
                else:
                    cell.font = styles['fail']['font']
                    cell.fill = styles['fail']['fill']
                cell.alignment = styles['data_center']['align']
            elif col_idx in [3, 8, 13]:  # 도메인/툴호출수/Reward
                cell.alignment = styles['data_center']['align']
            else:
                cell.alignment = styles['data']['align']
    
    # Freeze and filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    
    # Column widths
    ws.column_dimensions['A'].width = 7   # 결과
    ws.column_dimensions['B'].width = 28  # 모델
    ws.column_dimensions['C'].width = 10  # 도메인
    ws.column_dimensions['D'].width = 24  # 도메인 설명
    ws.column_dimensions['E'].width = 12  # Query ID
    ws.column_dimensions['F'].width = 48  # 질문
    ws.column_dimensions['G'].width = 34  # GT
    ws.column_dimensions['H'].width = 8   # 툴호출수
    ws.column_dimensions['I'].width = 28  # 툴목록
    ws.column_dimensions['J'].width = 42  # 모델응답
    ws.column_dimensions['K'].width = 26  # 실패분류
    ws.column_dimensions['L'].width = 18  # 근거
    ws.column_dimensions['M'].width = 10  # reward

    ws.freeze_panes = "A2"

def create_turns_raw_sheet(wb, turns_rows, styles):
    """턴 단위 원본(요청/응답/툴콜/툴결과). 요약/짤림 없이 그대로 저장."""
    ws = wb.create_sheet("턴_원본", 5)
    headers = ["RunID", "모델", "도메인", "TaskID", "Trial", "TurnIdx", "Role", "Content(원본)", "ToolCalls(JSON 원본)", "ToolResult(원본)"]
    ws.append(headers)
    for c in ws[1]:
        c.font = styles["header"]["font"]
        c.fill = styles["header"]["fill"]
        c.alignment = styles["header"]["align"]
        c.border = styles["header"]["border"]
    for row in turns_rows:
        ws.append(row)
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(r, c)
            cell.border = styles["data"]["border"]
            cell.alignment = styles["data"]["align"] if c not in [6] else styles["data_center"]["align"]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 6
    ws.column_dimensions["F"].width = 7
    ws.column_dimensions["G"].width = 8
    ws.column_dimensions["H"].width = 70
    ws.column_dimensions["I"].width = 60
    ws.column_dimensions["J"].width = 60
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 60


def create_case_summary_sheet(wb, runs, styles):
    """
    사람이 바로 읽히는 '케이스_요약' 시트.
    - 요청/GT/모델응답/툴호출/판정근거(왜 맞음/왜 틀림)를 한 행에 배치
    """
    ws = wb.create_sheet("케이스_요약", 1)

    title = "케이스 요약 (요청/정답기준/모델행동/판정근거)"
    ws.append([title])
    ws.merge_cells("A1:N1")
    ws["A1"].font = styles["title"]["font"]
    ws["A1"].alignment = styles["title"]["align"]
    ws.row_dimensions[1].height = 22

    ws.append(
        [
            "읽는 법: TAU2의 '정답(GT)'은 단일 텍스트가 아니라, (1) 필수 툴 호출/절차(Action) (2) DB 상태 변화(ENV/DB) (3) 커뮤니케이션 요건을 만족했는지로 판정됩니다.",
        ]
    )
    ws.merge_cells("A2:N2")
    ws["A2"].alignment = styles["data"]["align"]
    ws.row_dimensions[2].height = 36

    headers = [
        "RunID",
        "모델",
        "도메인",
        "TaskID",
        "Trial",
        "요청(원문)",
        "GT 요약(필수액션)",
        "모델 툴호출(요약)",
        "모델 최종응답(원문)",
        "결과",
        "Reward",
        "Breakdown(DB/ACTION/COMM)",
        "왜 맞음/왜 틀림(근거)",
        "참고(턴_원본 RunID)",
    ]
    ws.append(headers)
    header_row = ws.max_row
    for c in ws[header_row]:
        c.font = styles["header"]["font"]
        c.fill = styles["header"]["fill"]
        c.alignment = styles["header"]["align"]
        c.border = styles["header"]["border"]

    for run in runs:
        rb = run.get("RewardBreakdown") or {}
        breakdown = f'DB={rb.get("DB","")}, ACTION={rb.get("ACTION","")}, COMM={rb.get("COMMUNICATE","")}'

        # 판정근거(사람용)
        if run.get("Pass") == 1:
            why = "reward=1.0. 필수 액션(action_checks)이 모두 충족되고 DB/커뮤니케이션 체크가 통과된 케이스입니다."
        else:
            why_parts = []
            if run.get("Termination") and "too_many_errors" in str(run.get("Termination")):
                why_parts.append("too_many_errors로 조기 종료(실행 중 오류 누적).")
            if run.get("MissingRequiredActions"):
                why_parts.append(f"필수 액션 미충족: {run['MissingRequiredActions']}")
            if run.get("ActionMismatchCount") is not None:
                why_parts.append(f"action_checks 불일치 {run['ActionMismatchCount']}건.")
            if not why_parts:
                why_parts.append("reward_info 기준 미충족(세부 근거는 런_원본의 RewardBreakdown/툴로그 확인).")
            why = " ".join(why_parts)

        row = [
            run.get("RunID", ""),
            run.get("ModelLabel", ""),
            run.get("Domain", ""),
            run.get("TaskID", ""),
            run.get("Trial", 0),
            run.get("UserRequestRaw", ""),
            run.get("GTSummary", ""),
            run.get("ToolNames", ""),
            run.get("AgentFinalRaw", ""),
            "PASS" if run.get("Pass") == 1 else "FAIL",
            run.get("Reward", 0.0),
            breakdown,
            why,
            run.get("RunID", ""),
        ]
        ws.append(row)
        r = ws.max_row

        # 도메인 설명은 다른 시트에 숨겨져 있으니 여기서는 직접 텍스트로 넣지 않고, 가독성 확보를 위해 폭/줄바꿈으로 처리
        # 스타일
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(r, col_idx)
            cell.border = styles["data"]["border"]
            if col_idx in [5, 10, 11]:
                cell.alignment = styles["data_center"]["align"]
            else:
                cell.alignment = styles["data"]["align"]
        # 결과 컬럼 색 약하게
        result_cell = ws.cell(r, 10)
        if run.get("Pass") == 1:
            result_cell.fill = styles["pass"]["fill"]
            result_cell.font = styles["pass"]["font"]
        else:
            result_cell.fill = styles["fail"]["fill"]
            result_cell.font = styles["fail"]["font"]

        ws.row_dimensions[r].height = 84

    ws.freeze_panes = f"A{header_row+1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{ws.max_row}"

    # Column widths (읽기 최적화)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 6
    ws.column_dimensions["F"].width = 52
    ws.column_dimensions["G"].width = 30
    ws.column_dimensions["H"].width = 26
    ws.column_dimensions["I"].width = 52
    ws.column_dimensions["J"].width = 7
    ws.column_dimensions["K"].width = 8
    ws.column_dimensions["L"].width = 20
    ws.column_dimensions["M"].width = 44
    ws.column_dimensions["N"].width = 26


# 평가 대상 모델(표 표시 순서 고정)
# - 다른 스크립트(generate_reports.py)에서도 동일한 순서를 재사용하기 위해 상수로 분리
LLM_TO_LABEL: dict[str, str] = {
    "openrouter/meta-llama/llama-3.3-70b-instruct": "llama-3.3-70b-instruct-FC",
    "openrouter/mistralai/mistral-small-3.2-24b-instruct": "mistral-small-3.2-24b-instruct-FC",
    "openrouter/qwen/qwen3-32b": "qwen3-32b-FC",
    "openrouter/qwen/qwen3-14b": "qwen3-14b-FC",
    "openrouter/qwen/qwen3-next-80b-a3b-instruct": "qwen3-next-80b-a3b-instruct-FC",
}

def _find_default_base_dir() -> Path | None:
    base_dirs = [Path("data/simulations"), Path("data/tau2/simulations")]
    return next((d for d in base_dirs if d.exists()), None)


def generate_report(
    *,
    output_path: Path,
    model_filter: str | None = None,
    base_dir: Path | None = None,
    models_mapping_override: dict[str, str] | None = None,
) -> None:
    """
    리포트 생성 엔트리포인트(재사용 가능).
    - output_path: 저장할 xlsx 경로
    - model_filter: 특정 LLM 문자열(예: openrouter/... )만 포함하고 싶을 때
    - base_dir: 결과 json 폴더(기본: data/simulations 또는 data/tau2/simulations 자동 탐색)
    """
    base_dir = base_dir or _find_default_base_dir()
    if not base_dir:
        raise RuntimeError("Results directory not found. (expected data/simulations or data/tau2/simulations)")

    # 모델 매핑은 {llm_string: label} 형태로 유지
    models_mapping = dict(models_mapping_override or LLM_TO_LABEL)
    if model_filter:
        models_mapping = {k: v for k, v in models_mapping.items() if k == model_filter}

    domains = ["retail", "airline", "telecom"]
    all_logs = []
    all_turns = []

    # GT(정답) 성격: TAU2는 단일 정답 문자열이 아니라 "필수 액션/체크"의 집합
    gt_map: dict[tuple[str, str], str] = {}

    # 결과 JSON 파일 전체를 스캔해서 모델/도메인/시뮬레이션을 추출
    runs: list[dict] = []
    turns_rows: list[list] = []
    # task -> (user scenario raw, gt raw)
    task_meta: dict[tuple[str, str], dict] = {}

    for file_path in sorted(base_dir.glob("*.json")):
        try:
            data = json.loads(file_path.read_text())
        except Exception:
            continue

        info = data.get("info") or {}
        agent_llm = (((info.get("agent_info") or {}).get("llm")) or "").strip()
        domain = (((info.get("environment_info") or {}).get("domain_name")) or "").strip()
        if not agent_llm or not domain:
            continue

        # 관심 모델만 포함 (그 외는 제외)
        if agent_llm not in models_mapping:
            continue
        display_name = models_mapping[agent_llm]

        # GT 맵 구성 (task_id -> required action tool names)
        for t in data.get("tasks", []) or []:
            tid = str(t.get("id"))
            crit = (t.get("evaluation_criteria") or {})
            actions = crit.get("actions") or []
            # 원본 GT: actions + env_assertions(원문 그대로)
            gt_raw = json.dumps(
                {
                    "actions": actions,
                    "env_assertions": (crit.get("env_assertions") or []),
                },
                ensure_ascii=False,
            )
            req_tools = [a.get("name") for a in actions if a.get("name")]
            if req_tools:
                gt_map[(domain, tid)] = "required_tools: " + ", ".join(req_tools)
            task_meta[(domain, tid)] = {
                "gt_raw": gt_raw,
                "user_scenario_raw": json.dumps(((t.get("user_scenario") or {}).get("instructions") or {}), ensure_ascii=False),
                "ticket": (t.get("ticket") or ""),
            }

        for sim in data.get("simulations", []) or []:
            task_id = str(sim.get("task_id", "N/A"))
            trial = sim.get("trial")
            if trial is None:
                trial = sim.get("info_trial_num")
            if trial is None:
                trial = 0

            reward = sim.get("reward")
            if reward is None:
                reward = ((sim.get("reward_info") or {}).get("reward"))
            if reward is None:
                reward = 0.0

            reward_info = sim.get("reward_info") or {}
            reward_breakdown = reward_info.get("reward_breakdown") or {}
            action_checks = reward_info.get("action_checks") or []
            # 실패 원인(원문 기반)
            failed_env_assertions = _extract_failed_env_assertions(reward_info)
            action_mismatches = _extract_action_mismatches(reward_info)
            reward_basis = reward_info.get("reward_basis") or []
            if isinstance(reward_basis, list):
                reward_basis_norm = [ _normalize_reward_key(x) for x in reward_basis ]
            else:
                reward_basis_norm = []

            rb_env = _get_rb_value(reward_breakdown, "ENV_ASSERTION")
            rb_action = _get_rb_value(reward_breakdown, "ACTION")
            rb_db = _get_rb_value(reward_breakdown, "DB")
            rb_comm = _get_rb_value(reward_breakdown, "COMMUNICATE")
            rb_nl = _get_rb_value(reward_breakdown, "NL_ASSERTION")

            # TAU2 success 기준: reward가 1.0(±1e-6)
            is_pass = 1 if abs(float(reward) - 1.0) <= 1e-6 else 0

            run_id = f"{display_name}_{domain}_T{task_id}_trial{trial}"

            all_logs.append(
                {
                    "Model": display_name,
                    "Domain": domain,
                    "TaskID": task_id,
                    "Trial": int(trial),
                    "Pass": is_pass,
                    "Reward": float(reward),
                    "RewardBreakdown": reward_breakdown,
                    "Termination": sim.get("termination_reason", "N/A"),
                }
            )

            messages = sim.get("messages", []) or []
            tool_args_err_cnt, tool_args_err_summary, tool_args_errs_raw = _extract_tool_args_json_errors(messages)
            # run-level aggregation for tool count/names + raw request/gt/agent final
            tool_names: list[str] = []
            tool_count = 0
            first_user = ""
            agent_final = ""
            tool_calls_raw_all: list[dict] = []
            tool_results_raw_all: list[str] = []
            for idx, msg in enumerate(messages):
                role = msg.get("role", "")
                content = msg.get("content", "")
                tool_calls = msg.get("tool_calls", []) or []
                tool_info = ""
                if tool_calls:
                    tool_info = "; ".join(
                        [
                            f"{tc.get('name')}({json.dumps(tc.get('arguments', {}), ensure_ascii=False)})"
                            for tc in tool_calls
                            if tc.get("name")
                        ]
                    )
                    tool_count += len(tool_calls)
                    for tc in tool_calls:
                        n = tc.get("name")
                        if n:
                            tool_names.append(n)
                        tool_calls_raw_all.append(tc)

                if not first_user and role == "user":
                    first_user = content or ""
                if role == "assistant":
                    # 마지막 assistant 원본: 텍스트가 없고 tool_calls만 있으면 tool_calls를 텍스트로 표시
                    if content:
                        agent_final = content
                    elif tool_calls:
                        agent_final = json.dumps(tool_calls, ensure_ascii=False)
                    else:
                        agent_final = agent_final
                if role == "tool" and content:
                    tool_results_raw_all.append(content)

                all_turns.append(
                    {
                        "RunID": run_id,
                        "Turn": idx,
                        "Speaker": role,
                        "Message": content,
                        "Tool_Called": tool_info,
                        "Tool_Response": content[:200] + "..." if role == "tool" and len(content) > 200 else (content if role == "tool" else ""),
                    }
                )

                # 턴 원본 시트 row
                # 직관형 턴 시트용: kind/tool_name/tool_args/text/tool_result로 분리
                kind = "TEXT"
                tool_names_join = ""
                tool_args_join = ""
                tool_calls_json = json.dumps(tool_calls, ensure_ascii=False) if tool_calls else ""
                if role == "tool":
                    kind = "TOOL_RESULT"
                elif role == "assistant" and tool_calls:
                    kind = "TOOL_CALL"
                    names = []
                    args = []
                    for tc in tool_calls:
                        n = tc.get("name")
                        if n:
                            names.append(n)
                        args.append(json.dumps(tc.get("arguments", {}), ensure_ascii=False))
                    tool_names_join = "; ".join(names)
                    tool_args_join = "\n".join(args)
                turns_rows.append(
                    [
                        run_id,
                        display_name,
                        domain,
                        task_id,
                        int(trial),
                        "PASS" if is_pass == 1 else "FAIL",
                        idx,
                        role,
                        kind,
                        tool_names_join,
                        tool_args_join,
                        content if role != "tool" else "",
                        content if role == "tool" else "",
                        tool_calls_json,
                    ]
                )

            meta = task_meta.get((domain, str(task_id)), {})
            # 요청(원문)은 "시나리오 instructions JSON"을 기본으로 사용(첫 발화는 별도 컬럼으로)
            user_scenario_raw = meta.get("user_scenario_raw", "")
            user_request_raw = user_scenario_raw or first_user
            gt_raw = meta.get("gt_raw", "")
            gt_summary = gt_map.get((domain, str(task_id)), "N/A")
            # 필수 툴 리스트(GT)
            required_tools: list[str] = []
            try:
                gt_obj = json.loads(gt_raw) if gt_raw else {}
                if isinstance(gt_obj, dict):
                    for a in (gt_obj.get("actions") or []):
                        if isinstance(a, dict) and a.get("name"):
                            required_tools.append(a["name"])
            except Exception:
                required_tools = []
            called_tools = sorted({n for n in tool_names if n})
            missing_tools = sorted(set(required_tools) - set(called_tools))

            # 실패 근거(필수 액션 미충족 등) 계산
            mismatches = [a for a in action_checks if a and (a.get("action_match") is False)]
            mismatch_count = len(mismatches)
            missing_actions = []
            for a in mismatches[:8]:
                act = (a.get("action") or {})
                name = act.get("name")
                if name:
                    missing_actions.append(name)
            missing_actions_str = ", ".join(missing_actions)

            runs.append(
                {
                    "RunID": run_id,
                    "ModelLabel": display_name,
                    "AgentLLM": agent_llm,
                    "Domain": domain,
                    "TaskID": task_id,
                    "Trial": int(trial),
                    "Pass": is_pass,
                    "Reward": float(reward),
                    "RewardBreakdownJSON": json.dumps(reward_breakdown, ensure_ascii=False),
                    "RewardBreakdown": reward_breakdown,
                    "RewardBasisRaw": json.dumps(reward_basis_norm, ensure_ascii=False),
                    "RB_ENV_ASSERTION": rb_env,
                    "RB_ACTION": rb_action,
                    "RB_DB": rb_db,
                    "RB_COMMUNICATE": rb_comm,
                    "RB_NL_ASSERTION": rb_nl,
                    "Termination": sim.get("termination_reason", "N/A"),
                    "ToolCallCount": tool_count,
                    "ToolNames": ", ".join(tool_names),
                    "ToolCallsRaw": json.dumps(tool_calls_raw_all, ensure_ascii=False),
                    "ToolResultsRaw": "\n\n---\n\n".join(tool_results_raw_all),
                    "UserRequestRaw": user_request_raw,
                    "UserScenarioRaw": user_scenario_raw,
                    "UserFirstUtterance": first_user,
                    "GTRaw": gt_raw,
                    "GTSummary": gt_summary,
                    "AgentFinalRaw": agent_final,
                    "ActionChecksRaw": json.dumps(action_checks, ensure_ascii=False),
                    "ActionMismatchCount": mismatch_count,
                    "MissingRequiredActions": missing_actions_str,
                    "RequiredTools": sorted(set(required_tools)),
                    "CalledTools": called_tools,
                    "MissingTools": missing_tools,
                    "FailedEnvAssertions": failed_env_assertions,
                    "FailedEnvAssertionCount": len(failed_env_assertions),
                    "ActionMismatches": action_mismatches,
                    "ToolArgsJSONErrorCount": tool_args_err_cnt,
                    "ToolArgsJSONErrorSummary": tool_args_err_summary,
                    "ToolArgsJSONErrorsRaw": tool_args_errs_raw,
                }
            )

    wb = Workbook()
    wb.remove(wb.active)
    styles = setup_styles()
    # 엑셀 열 때 수식 재계산(캐시값 NULL 방지)
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass
    
    # 도메인 설명 시트(다른 시트에서 VLOOKUP으로 참조)
    ws_dom = wb.create_sheet("도메인_설명", 0)
    ws_dom.append(["domain", "표시명", "설명"])
    for c in ws_dom[1]:
        c.font = styles["header"]["font"]
        c.fill = styles["header"]["fill"]
        c.alignment = styles["header"]["align"]
        c.border = styles["header"]["border"]
    rows = [
        ["retail", "Retail", "온라인 쇼핑 주문/반품/교환/배송지/결제 변경 등 고객센터 시나리오"],
        ["airline", "Airline", "항공 예약/변경/좌석/마일리지 등 고객센터 시나리오"],
        ["telecom", "Telecom", "요금제/장애/청구/개통 등 통신 고객센터 시나리오"],
    ]
    for r in rows:
        ws_dom.append(r)
    ws_dom.column_dimensions["A"].width = 10
    ws_dom.column_dimensions["B"].width = 10
    ws_dom.column_dimensions["C"].width = 60

    # ===== 3-sheet layout (visible) =====
    create_summary_sheet(wb, models_mapping, domains, styles)     # 요약(랭킹+매트릭스+Glossary)
    create_runs_sheet(wb, runs, styles)                           # 런(케이스 단위, 원본 + 실패사유)
    create_turns_sheet(wb, turns_rows, styles)                    # 대화(원문 + TOOL_CALL/RESULT)

    # ===== helper sheets (hidden) =====
    create_task_summary_sheet(wb, all_logs, models_mapping, domains, styles)  # Pass^k 계산용

    # 시트 수 줄이기: helper 시트 숨김
    for name in ["Task별_집계", "도메인_설명"]:
        if name in wb.sheetnames:
            wb[name].sheet_state = "hidden"
    
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    # macOS Excel 경고(외부 데이터 연결/신뢰) 오탐을 줄이기 위해 xattr 제거
    # - 실제로 외부 연결은 포함되어 있지 않음(패키지 검사 기준).
    try:
        if platform.system().lower() == "darwin":
            # 가장 확실: 모든 xattr 제거
            subprocess.run(["xattr", "-c", str(output_path)], check=False, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    except Exception:
        pass
    # 로그 출력은 사용자가 CLI에서 쓸 때만 의미가 있으므로, 호출자가 선택적으로 처리


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--model-filter", type=str, default=None, help="LLM 문자열(예: openrouter/...)로 단일 모델만 리포트 생성")
    parser.add_argument("--output", type=str, default="tau2_evaluation_report.xlsx", help="출력 xlsx 경로(기본: tau2_evaluation_report.xlsx)")
    parser.add_argument("--input-dir", type=str, default=None, help="시뮬레이션 json 폴더(기본: data/simulations 자동 탐색)")
    args = parser.parse_args()

    base_dir = Path(args.input_dir) if args.input_dir else None
    out = Path(args.output)
    generate_report(output_path=out, model_filter=args.model_filter, base_dir=base_dir)

    print(f"\nReport generated: {out}")
    print(f"  - Visible sheets: 요약, 런, 턴 (helper는 숨김)")
    print(f"  - Data source: {str(base_dir or _find_default_base_dir())}")

if __name__ == "__main__":
    main()
